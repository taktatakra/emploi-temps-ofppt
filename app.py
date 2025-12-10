#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Streamlit app : Gestionnaire d'Emploi du Temps (Mode libre / neutre)
# - Mode libre : vous pouvez saisir un libellé de mois libre et créer/modifier des semaines (S1, S2, ...)
# - Si vous importez un fichier Excel, ses onglets/mois sont disponibles mais aucune présélection n'est faite.
# - Les semaines peuvent être importées depuis l'onglet détecté ou créées manuellement par l'utilisateur.
#
# Usage: streamlit run app.py
#
# Dépendances: streamlit, pandas, openpyxl, plotly
# Placez Logo_ofppt.png dans le répertoire pour l'utiliser localement.

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime, timedelta, date
import copy
import os
import re
import base64

# --- Page config ---
st.set_page_config(
    page_title="Gestionnaire d'Emploi du Temps - OFPPT (Mode libre)",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- SESSION STATE defaults ---
if 'raw_data' not in st.session_state:
    st.session_state['raw_data'] = None
if 'resolved_data' not in st.session_state:
    st.session_state['resolved_data'] = None
if 'conflits_log' not in st.session_state:
    st.session_state['conflits_log'] = pd.DataFrame()
if 'niveau_global' not in st.session_state:
    st.session_state['niveau_global'] = "1ère Année"
if 'force_25_to_26' not in st.session_state:
    st.session_state['force_25_to_26'] = True
# custom week ranges keyed by month label (string)
if 'custom_week_ranges' not in st.session_state:
    st.session_state['custom_week_ranges'] = {}

# --- STYLE ---
st.markdown("""
<style>
    .main-header { background: linear-gradient(135deg,#1e5631 0%,#2d8659 50%,#1e5631 100%); padding: 2.0rem; border-radius: 12px; margin-bottom: 1.5rem; color: white; text-align:center; }
    .ofppt-title { font-size: 2.4rem; font-weight: bold; margin-bottom: 0.3rem; }
    .ofppt-subtitle { font-size: 1.0rem; margin-bottom: 0.5rem; opacity: 0.95; }
    .section-header { font-size: 1.3rem; color: #1e5631; font-weight: bold; margin: 1.2rem 0 0.8rem 0; padding-bottom: 0.4rem; border-bottom: 2px solid #2d8659; }
    .metric-card { background: white; padding: 0.9rem; border-radius: 8px; border-left: 4px solid #2d8659; text-align:center; box-shadow:0 1px 6px rgba(0,0,0,0.04); }
</style>
""", unsafe_allow_html=True)

# --- CONSTANTS ---
LOGO_FILE_NAME = "Logo_ofppt.png"
LOGO_URL = "https://www.ofppt.ma/sites/default/files/logo.png"
LOGO_WIDTH_PIXELS = 70
LOGO_HEIGHT_PIXELS = 70

FALLBACK_SEMAINES = ['S1', 'S2', 'S3', 'S4']

JOURS = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
CRENEAUX_JOUR = ['AM1', 'AM2', 'PM1', 'PM2']

HORAIRES = {
    'AM1': '08H30-11H00',
    'AM2': '11H00-13H30',
    'PM1': '13H30-16H00',
    'PM2': '16H00-18H30'
}

SLOT_DURATIONS = {'AM1': 2.5, 'AM2': 2.5, 'PM1': 2.5, 'PM2': 2.5}

MONTH_NAMES = {
    'Novembre': 'Novembre','Decembre': 'Décembre','Janvier':'Janvier','Février':'Février',
    'Mars':'Mars','Avril':'Avril','Mai':'Mai','Juin':'Juin','Juillet':'Juillet',
    'Aout':'Août','Août':'Août','Septembre':'Septembre','Octobre':'Octobre'
}
MONTH_ORDER = list(MONTH_NAMES.values())
IGNORED_SHEETS = ['Groupes', 'Feuil1', 'Sheet1']

MONTH_TO_NUMBER = {
    'Janvier':1,'Février':2,'Mars':3,'Avril':4,'Mai':5,'Juin':6,'Juillet':7,'Août':8,'Aout':8,
    'Septembre':9,'Octobre':10,'Novembre':11,'Décembre':12
}

# Example holidays (adjust as needed)
HOLIDAYS = [
    {'date': datetime(2025,11,6).date(), 'label': 'La Marche Verte'},
    {'date': datetime(2025,11,18).date(), 'label': "Fête de l'Indépendance"},
    {'date': datetime(2026,1,1).date(), 'label': 'Nouvel an'},
    {'date': datetime(2026,1,11).date(), 'label': "Manifeste de l'independence"},
    {'date': datetime(2026,1,14).date(), 'label': 'Nouvel an Amazigh'},
    {'date': datetime(2026,5,1).date(), 'label': 'Fête du travail'},
]

HOLIDAY_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
HOLIDAY_FONT = Font(bold=True, color="000000")

# --- UTIL ---
def get_logo_src():
    if os.path.exists(LOGO_FILE_NAME):
        try:
            with open(LOGO_FILE_NAME, "rb") as f:
                return f"data:image/png;base64,{base64.b64encode(f.read()).decode()}"
        except Exception:
            return LOGO_URL
    return LOGO_URL

ARROW_RE = re.compile(r'\s*(?:→|->|–|-)\s*')
DATE_FORMATS = ["%d/%m/%Y","%d/%m/%y","%d %b %Y","%d %B %Y","%d %b %y","%d %B %y","%Y-%m-%d","%d.%m.%Y"]

def try_parse_date(s):
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    s = re.sub(r'[^\w\s\-/\.]', ' ', s).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    parts = re.split(r'[ \-\\/\.]+', s)
    if len(parts) >= 3:
        try:
            d = int(parts[0]); m = int(parts[1]); y = int(parts[2])
            if y < 100: y += 2000
            return datetime(y, m, d).date()
        except Exception:
            pass
    return None

def parse_date_range_cell(cell):
    if not cell or not isinstance(cell, str):
        return (None, None)
    if '→' in cell or '->' in cell or '–' in cell or '-' in cell:
        parts = ARROW_RE.split(cell)
        if len(parts) >= 2:
            d1 = try_parse_date(parts[0])
            d2 = try_parse_date(parts[1])
            return (d1, d2)
    return (None, None)

def sanitize_sheet_title(s, max_len=31):
    if s is None:
        return "Sheet1"
    s = str(s)
    s = re.sub(r'[:\\\/\?\*\[\]]', '_', s)
    s = s.strip() or "sheet"
    if len(s) > max_len:
        s = s[:max_len]
    return s

def is_holiday(day_date):
    if day_date is None:
        return None
    if isinstance(day_date, datetime):
        d = day_date.date()
    else:
        d = day_date
    for h in HOLIDAYS:
        if h['date'] == d:
            return h['label']
    return None

def day_date(week_start, offset_days):
    if week_start is None:
        return None
    return week_start + timedelta(days=offset_days)

# --- PARSING of uploaded Excel ---
def find_header_row(df):
    for idx, row in df.iterrows():
        vals = [str(x).strip() for x in row.values if pd.notna(x)]
        lowvals = [v.lower() for v in vals]
        if any(v in ('formateur','form') for v in lowvals) and any(c in vals for c in ['AM1','AM2','PM1','PM2']):
            return idx
    return None

@st.cache_data(show_spinner=False)
def parse_schedule_sheet(df, sheet_name):
    month_name = sheet_name.replace('Planning_', '').strip()
    month_label = month_name if month_name else sheet_name

    header_idx = find_header_row(df)
    if header_idx is None:
        return None

    header_row = df.iloc[header_idx]
    search_top = max(0, header_idx - 10)
    found_map = {}
    for ridx in range(search_top, header_idx):
        row = df.iloc[ridx].astype(str).tolist()
        for cidx, cell in enumerate(row):
            txt = str(cell).strip()
            if not txt:
                continue
            a,b = parse_date_range_cell(txt)
            if a and b and cidx not in found_map:
                found_map[cidx] = (txt, a, b)
    ordered = [found_map[k] for k in sorted(found_map.keys())] if found_map else []
    if ordered:
        semaines = [it[0] for it in ordered]
        week_ranges = {it[0]: {'start': it[1], 'end': it[2]} for it in ordered}
    else:
        semaines = []
        week_ranges = {}

    col_form = -1
    for i, val in enumerate(header_row):
        if pd.notna(val) and str(val).strip().lower() in ('formateur','form'):
            col_form = i
            break
    if col_form == -1:
        return None
    col_salle = col_form + 1
    col_start = col_salle + 1

    df_data = df.iloc[header_idx+1:].reset_index(drop=True)

    schedule = {}
    groupes = set()
    salles = set()
    col_map = {}
    cur = col_start
    for s in semaines:
        for j in JOURS:
            for c in CRENEAUX_JOUR:
                col_map[f"{s}-{j}-{c}"] = cur
                cur += 1

    for _, row in df_data.iterrows():
        form = str(row.iloc[col_form]).strip() if col_form < len(row) else ''
        salle = str(row.iloc[col_salle]).strip() if col_salle < len(row) else ''
        if not form or form.lower() in ('nan','none',''):
            continue
        schedule.setdefault(form, {'salle': salle, 'slots': {}})
        if salle and salle.lower() not in ('nan','none',''):
            salles.add(salle)
        for key, ci in col_map.items():
            if ci < len(row):
                val = row.iloc[ci]
                if pd.notna(val) and str(val).strip():
                    grp = str(val).strip()
                    if grp and not grp.isdigit() and grp.lower() not in ('nan','none'):
                        schedule[form]['slots'][key] = (grp, salle)
                        groupes.add(grp)

    return {
        'month': month_label,
        'schedule': schedule,
        'formateurs': sorted(schedule.keys()),
        'groupes': sorted(list(groupes)),
        'salles': sorted(list(salles)),
        'semaines': semaines,
        'week_ranges': week_ranges,
        'header_idx': int(header_idx)
    }

@st.cache_data(show_spinner=False)
def process_uploaded_excel(uploaded_file):
    all_data = {}
    try:
        xls = pd.ExcelFile(uploaded_file)
        for sheet_name in xls.sheet_names:
            if sheet_name in IGNORED_SHEETS:
                continue
            df = pd.read_excel(xls, sheet_name=sheet_name, header=None, dtype=str)
            df = df.fillna('')
            parsed = parse_schedule_sheet(df, sheet_name)
            if parsed:
                all_data[parsed['month']] = parsed
        sorted_data = {m: all_data[m] for m in MONTH_ORDER if m in all_data}
        for k, v in all_data.items():
            if k not in sorted_data:
                sorted_data[k] = v
        return sorted_data
    except Exception as e:
        st.error(f"Erreur import: {e}")
        return {}

# --- CONFLICT RESOLUTION (full) ---
@st.cache_data(show_spinner=False)
def resolve_salle_conflits(all_data):
    resolved = copy.deepcopy(all_data)
    log = []
    all_salles = set()
    for month in resolved.values():
        all_salles.update(month['salles'])
    for month_name, month_data in resolved.items():
        schedule = month_data['schedule']
        semaines = month_data.get('semaines', FALLBACK_SEMAINES)
        HALF_DAY = [('AM1','AM2'), ('PM1','PM2')]
        for semaine in semaines:
            for jour in JOURS:
                for c1, c2 in HALF_DAY:
                    key1 = f"{semaine}-{jour}-{c1}"
                    key2 = f"{semaine}-{jour}-{c2}"
                    occ1 = set()
                    occ2 = set()
                    for f, fd in schedule.items():
                        s1 = fd['slots'].get(key1)
                        s2 = fd['slots'].get(key2)
                        if s1 and s1[1]:
                            occ1.add(s1[1])
                        if s2 and s2[1]:
                            occ2.add(s2[1])
                    libres = all_salles - occ1 - occ2
                    requests = []
                    for form, fd in schedule.items():
                        si1 = fd['slots'].get(key1, (None,None))
                        si2 = fd['slots'].get(key2, (None,None))
                        g1, g2 = si1[0], si2[0]
                        if g1 or g2:
                            requests.append({'formateur': form, 'g1': g1, 'g2': g2, 'pref': fd['salle']})
                    occupied = set()
                    for req in requests:
                        f = req['formateur']
                        pref = req['pref']
                        assigned = None
                        if pref and pref not in occupied:
                            assigned = pref
                        else:
                            candidates = libres - occupied
                            candidates = [s for s in candidates if not any(x in s.lower() for x in ['info','ent'])]
                            if candidates:
                                candidates.sort()
                                assigned = candidates[0]
                                for creneau, grp in [(c1, req['g1']), (c2, req['g2'])]:
                                    if grp:
                                        log.append({'Mois': month_name, 'Semaine': semaine, 'Jour_Creneau': f"{jour}-{creneau}", 'Heure': HORAIRES[creneau], 'Formateur': f, 'Groupe': grp, 'Salle_Initiale': pref, 'Salle_Attribuee': assigned})
                            else:
                                assigned = f"{pref or 'Aucune'} (CONFLIT NON RESOLU)"
                                for creneau, grp in [(c1, req['g1']), (c2, req['g2'])]:
                                    if grp:
                                        log.append({'Mois': month_name, 'Semaine': semaine, 'Jour_Creneau': f"{jour}-{creneau}", 'Heure': HORAIRES[creneau], 'Formateur': f, 'Groupe': grp, 'Salle_Initiale': pref, 'Salle_Attribuee': 'AUCUNE DISPO'})
                        if "CONFLIT NON RESOLU" not in assigned:
                            occupied.add(assigned)
                        if req['g1']:
                            resolved[month_name]['schedule'][f]['slots'][key1] = (req['g1'], assigned)
                        if req['g2']:
                            resolved[month_name]['schedule'][f]['slots'][key2] = (req['g2'], assigned)
    return resolved, pd.DataFrame(log)

# --- week helper (prioritize custom week_ranges supplied by user) ---
def get_week_start_from_label(mois_label, semaine_label, week_ranges):
    if week_ranges and semaine_label in week_ranges:
        return week_ranges[semaine_label]['start']
    mnum = MONTH_TO_NUMBER.get(mois_label)
    if mnum:
        year = 2026 if mnum <= 7 else 2025
        first_day = datetime(year, mnum, 1)
        days_until_monday = (7 - first_day.weekday()) % 7
        first_monday = first_day + timedelta(days=days_until_monday) if first_day.weekday() != 0 else first_day
        m = re.match(r'S(\d+)', str(semaine_label or ''), re.I)
        if m:
            idx = int(m.group(1)) - 1
            return (first_monday + timedelta(weeks=idx)).date()
    return None

# --- EXCEL export helpers (full functions) ---
def add_logo_if_exists(ws, cell='A1'):
    try:
        if os.path.exists(LOGO_FILE_NAME):
            img = Image(LOGO_FILE_NAME)
            img.width = LOGO_WIDTH_PIXELS
            img.height = LOGO_HEIGHT_PIXELS
            ws.add_image(img, cell)
    except Exception:
        pass

def excel_to_bytes(wb):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def clear_row_borders(ws, row_idx, start_col=1, end_col=5):
    empty_border = Border()
    for c in range(start_col, end_col + 1):
        try:
            ws.cell(row=row_idx, column=c).border = empty_border
        except Exception:
            pass

def clear_meta_borders(ws, meta_top_row=5, start_col=1, end_col=5):
    empty_border = Border()
    for r in range(1, meta_top_row + 1):
        for c in range(start_col, end_col + 1):
            try:
                ws.cell(row=r, column=c).border = empty_border
            except Exception:
                pass

def find_and_clear_signature_rows(ws, signature_text='Directeur EFP', start_col=1, end_col=5):
    try:
        sig_rows = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and signature_text.lower() in v.strip().lower():
                sig_rows.append(r)
        for r in sig_rows:
            clear_row_borders(ws, r, start_col, end_col)
            if r - 1 >= 1:
                clear_row_borders(ws, r - 1, start_col, end_col)
    except Exception:
        pass

def _apply_template_title(ws, title_text, heures_text, periode_text, left_meta, right_meta):
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_font = Font(bold=True, size=14, name='Calibri')
    meta_font_bold = Font(bold=True, size=10, name='Calibri')
    ws.merge_cells('B1:E2')
    ws['B1'] = title_text
    ws['B1'].font = title_font
    ws['B1'].alignment = center_align

    ws.merge_cells('B3:E3')
    ws['B3'] = heures_text
    ws['B3'].font = meta_font_bold
    ws['B3'].alignment = center_align

    ws.merge_cells('B4:E4')
    ws['B4'] = periode_text
    ws['B4'].font = meta_font_bold
    ws['B4'].alignment = center_align

    for idx, (cell, value) in enumerate(left_meta, start=5):
        ws[cell] = value
        ws[cell].font = meta_font_bold
        ws[cell].alignment = Alignment(horizontal='left', vertical='center')

    for idx, val in enumerate(right_meta, start=5):
        ws[f'E{idx}'] = val
        ws[f'E{idx}'].font = meta_font_bold
        ws[f'E{idx}'].alignment = Alignment(horizontal='right', vertical='center')

def _draw_table_borders(ws, start_row, end_row, start_col=1, end_col=5, meta_top_row=5):
    thin_side = Side(style='thin', color='000000')
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            try:
                ws.cell(row=r, column=c).border = border_all
            except Exception:
                pass

    for r in range(meta_top_row, end_row + 1):
        left_cell = ws.cell(row=r, column=start_col)
        try:
            existing = left_cell.border
            left_cell.border = Border(
                left=thin_side,
                right=existing.right if existing else thin_side,
                top=existing.top if existing else thin_side,
                bottom=existing.bottom if existing else thin_side
            )
        except Exception:
            left_cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for r in range(start_row, end_row + 1):
        right_cell = ws.cell(row=r, column=end_col)
        try:
            existing = right_cell.border
            right_cell.border = Border(
                left=existing.left if existing else thin_side,
                right=thin_side,
                top=existing.top if existing else thin_side,
                bottom=existing.bottom if existing else thin_side
            )
        except Exception:
            right_cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    try:
        for m in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = m.bounds
            if max_row >= start_row and min_row <= end_row and not (max_col < start_col or min_col > end_col):
                r1 = max(min_row, start_row)
                r2 = min(max_row, end_row)
                c1 = max(min_col, start_col)
                c2 = min(max_col, end_col)
                for rr in range(r1, r2 + 1):
                    for cc in range(c1, c2 + 1):
                        cell = ws.cell(row=rr, column=cc)
                        cell.border = border_all
    except Exception:
        pass

def copy_sheet(ws_src, ws_dest):
    from openpyxl.cell.cell import MergedCell
    import copy as _copy

    try:
        ws_dest.sheet_view.showGridLines = ws_src.sheet_view.showGridLines
    except Exception:
        pass

    try:
        for merged in ws_src.merged_cells.ranges:
            ws_dest.merge_cells(str(merged))
    except Exception:
        pass

    try:
        for col_letter, col_dim in ws_src.column_dimensions.items():
            try:
                ws_dest.column_dimensions[col_letter].width = col_dim.width
            except Exception:
                pass
    except Exception:
        pass

    try:
        for idx, row_dim in ws_src.row_dimensions.items():
            try:
                ws_dest.row_dimensions[idx].height = row_dim.height
            except Exception:
                pass
    except Exception:
        pass

    for row in ws_src.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            new = ws_dest.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            try:
                if getattr(cell, "has_style", False):
                    new.font = _copy.copy(cell.font)
                    new.border = _copy.copy(cell.border)
                    new.fill = _copy.copy(cell.fill)
                    new.number_format = cell.number_format
                    new.protection = _copy.copy(cell.protection)
                    new.alignment = _copy.copy(cell.alignment)
            except Exception:
                new.value = cell.value

    try:
        ws_dest.page_setup.orientation = ws_src.page_setup.orientation
        ws_dest.page_setup.paperSize = ws_src.page_setup.paperSize
        ws_dest.page_setup.fitToPage = ws_src.page_setup.fitToPage
        ws_dest.page_setup.fitToHeight = ws_src.page_setup.fitToHeight
        ws_dest.page_setup.fitToWidth = ws_src.page_setup.fitToWidth
    except Exception:
        pass

    add_logo_if_exists(ws_dest, 'A1')

    header_row = None
    try:
        for r in range(1, min(30, ws_dest.max_row) + 1):
            val = ws_dest.cell(row=r, column=1).value
            if isinstance(val, str) and val.strip().upper() == 'JOUR':
                header_row = r
                break
    except Exception:
        header_row = None

    if header_row:
        try:
            end_row = None
            for r in range(ws_dest.max_row, header_row - 1, -1):
                a_val = ws_dest.cell(row=r, column=1).value
                if (isinstance(a_val, str) and a_val.strip() in JOURS) or any([ws_dest.cell(row=r, column=c).value not in (None, '') for c in range(2,6)]):
                    end_row = r
                    break
            if end_row and end_row >= header_row:
                _draw_table_borders(ws_dest, header_row, end_row, 1, 5, meta_top_row=5)
            else:
                last_row = ws_dest.max_row
                _draw_table_borders(ws_dest, header_row, last_row, 1, 5, meta_top_row=5)
        except Exception:
            pass

    try:
        clear_meta_borders(ws_dest, meta_top_row=5, start_col=1, end_col=5)
    except Exception:
        pass
    try:
        find_and_clear_signature_rows(ws_dest, signature_text='Directeur EFP', start_col=1, end_col=5)
    except Exception:
        pass

def create_excel_formateur_semaine(formateur, data, semaine_label, mois_label, week_ranges, niveau="1ère Année", force_25_to_26=True):
    wb = openpyxl.Workbook()
    ws = wb.active
    raw_title = f"{formateur[:20]}-{mois_label[:10]}"
    ws.title = sanitize_sheet_title(raw_title)
    ws.sheet_view.showGridLines = False

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1

    add_logo_if_exists(ws, 'A1')

    try:
        start_dt = get_week_start_from_label(mois_label, semaine_label, week_ranges)
        end_dt = start_dt + timedelta(days=5) if start_dt else None
        periode_text = f"Date d'application: Du {start_dt.strftime('%d/%m/%Y')} au {end_dt.strftime('%d/%m/%Y')}" if start_dt and end_dt else ""
    except Exception:
        periode_text = ""

    title_text = 'EMPLOI DU TEMPS DE FORMATEUR : FORMATION HYBRIDE - V 1.0'

    heures_val_calc = compute_hours_for_formateur(data, semaine_label, mois_label, week_ranges)
    if force_25_to_26 and abs(heures_val_calc - 25.0) < 0.01:
        heures_val = 26.0
    else:
        heures_val = heures_val_calc

    heures_text = f'MASSE HORAIRE: {heures_val:.1f}H/SEMAINE'
    left_meta = [('A5', 'CFP TLRA/IFMLT'),
                 ('A6', f'Formateur: {formateur}'),
                 ('A7', f'Mois: {mois_label}'),
                 ('A8', 'Année de Formation: 2025/2026')]
    right_meta = ['', '', f'Niveau: {niveau}', '']
    _apply_template_title(ws, title_text, heures_text, periode_text, left_meta, right_meta)

    clear_meta_borders(ws, meta_top_row=5, start_col=1, end_col=5)

    header_row = 9
    header_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    headers = ['JOUR'] + [f"{c}\n{HORAIRES[c]}" for c in CRENEAUX_JOUR]
    for idx, txt in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=txt)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_thin
    ws.row_dimensions[header_row].height = 26

    row = header_row + 1
    week_start = get_week_start_from_label(mois_label, semaine_label, week_ranges)
    for j_idx, jour in enumerate(JOURS):
        ws.cell(row=row, column=1, value=jour).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=1).border = border_thin
        d = day_date(week_start, j_idx)
        holiday_label = is_holiday(d) if d else None
        if holiday_label:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            cell = ws.cell(row=row, column=2, value=holiday_label)
            cell.font = HOLIDAY_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = HOLIDAY_FILL
            for c in range(2,6):
                ws.cell(row=row, column=c).border = border_thin
        else:
            for ci, creneau in enumerate(CRENEAUX_JOUR, start=2):
                key = f"{semaine_label}-{jour}-{creneau}"
                slot = data['slots'].get(key, ('',''))
                grp, salle = slot
                text = f"{grp}\n{salle}" if grp and salle else ""
                cell = ws.cell(row=row, column=ci, value=text)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(size=10, bold=True)
                cell.border = border_thin
        ws.row_dimensions[row].height = 28
        row += 1

    _draw_table_borders(ws, header_row, row-1, 1, 5, meta_top_row=5)
    clear_meta_borders(ws, meta_top_row=5, start_col=1, end_col=5)

    sig_row = row + 1
    ws.cell(row=sig_row, column=1, value='Directeur EFP').font = Font(size=10, bold=True)
    ws.cell(row=sig_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    try:
        clear_row_borders(ws, sig_row - 1, 1, 5)
    except Exception:
        pass
    try:
        clear_row_borders(ws, sig_row, 1, 5)
    except Exception:
        pass

    ws.column_dimensions['A'].width = 18
    for col in ['B','C','D','E']:
        ws.column_dimensions[col].width = 20

    return wb

def create_excel_groupe_semaine(groupe, schedule_data, semaine_label, mois_label, week_ranges, niveau="1ère Année"):
    wb = openpyxl.Workbook()
    ws = wb.active
    raw_title = f"{groupe[:20]}-{mois_label[:10]}"
    ws.title = sanitize_sheet_title(raw_title)
    ws.sheet_view.showGridLines = False

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1

    add_logo_if_exists(ws, 'A1')

    try:
        start_dt = get_week_start_from_label(mois_label, semaine_label, week_ranges)
        end_dt = start_dt + timedelta(days=5) if start_dt else None
        periode_text = f"Date d'application: Du {start_dt.strftime('%d/%m/%Y')} au {end_dt.strftime('%d/%m/%Y')}" if start_dt and end_dt else ""
    except Exception:
        periode_text = ""

    title_text = 'EMPLOI DU TEMPS PAR GROUPE : FORMATION HYBRIDE - V 1.0'
    heures_val = compute_hours_for_groupe(schedule_data, groupe, semaine_label, mois_label, week_ranges)
    heures_text = f'MASSE HORAIRE: {heures_val:.1f}H/SEMAINE'
    left_meta = [('A5', 'CFP TLRA/IFMLT'),
                 ('A6', f'Groupe: {groupe}'),
                 ('A7', f'Mois: {mois_label}'),
                 ('A8', 'Année de Formation: 2025/2026')]
    right_meta = ['', '', f'Niveau: {niveau}', '']
    _apply_template_title(ws, title_text, heures_text, periode_text, left_meta, right_meta)

    clear_meta_borders(ws, meta_top_row=5, start_col=1, end_col=5)

    header_row = 9
    header_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    headers = ['JOUR'] + [f"{c}\n{HORAIRES[c]}" for c in CRENEAUX_JOUR]
    for idx, txt in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=txt)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_thin
    ws.row_dimensions[header_row].height = 26

    row = header_row + 1
    week_start = get_week_start_from_label(mois_label, semaine_label, week_ranges)
    for j_idx, jour in enumerate(JOURS):
        ws.cell(row=row, column=1, value=jour).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        d = day_date(week_start, j_idx)
        holiday_label = is_holiday(d) if d else None
        if holiday_label:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            cell = ws.cell(row=row, column=2, value=holiday_label)
            cell.font = HOLIDAY_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = HOLIDAY_FILL
            for c in range(2,6):
                ws.cell(row=row, column=c).border = border_thin
        else:
            for ci, creneau in enumerate(CRENEAUX_JOUR, start=2):
                key = f"{semaine_label}-{jour}-{creneau}"
                info = ""
                for f, fd in schedule_data.items():
                    s = fd['slots'].get(key)
                    if s and s[0] == groupe:
                        info = f"{f}\n{s[1].replace(' (CONFLIT NON RESOLU)',' (Conflit)')}"
                        break
                cell = ws.cell(row=row, column=ci, value=info)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(size=10, bold=True)
                cell.border = border_thin
        ws.row_dimensions[row].height = 28
        row += 1

    _draw_table_borders(ws, header_row, row-1, 1, 5, meta_top_row=5)
    clear_meta_borders(ws, meta_top_row=5, start_col=1, end_col=5)

    sig_row = row + 1
    ws.cell(row=sig_row, column=1, value='Directeur EFP').font = Font(size=10, bold=True)
    ws.cell(row=sig_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    try:
        clear_row_borders(ws, sig_row - 1, 1, 5)
    except Exception:
        pass
    try:
        clear_row_borders(ws, sig_row, 1, 5)
    except Exception:
        pass

    ws.column_dimensions['A'].width = 18
    for col in ['B','C','D','E']:
        ws.column_dimensions[col].width = 20

    return wb

# --- compute hours ---
def compute_hours_for_formateur(formateur_data, semaine_label, mois_label, week_ranges):
    heures = 0.0
    week_start = get_week_start_from_label(mois_label, semaine_label, week_ranges)
    for jour_idx, jour in enumerate(JOURS):
        day_dt = day_date(week_start, jour_idx)
        if day_dt and is_holiday(day_dt):
            continue
        for c in CRENEAUX_JOUR:
            slot_key = f"{semaine_label}-{jour}-{c}"
            if slot_key in formateur_data.get('slots', {}):
                heures += SLOT_DURATIONS.get(c, 0)
    return heures

def compute_hours_for_groupe(schedule_data, groupe, semaine_label, mois_label, week_ranges):
    heures = 0.0
    week_start = get_week_start_from_label(mois_label, semaine_label, week_ranges)
    for jour_idx, jour in enumerate(JOURS):
        day_dt = day_date(week_start, jour_idx)
        if day_dt and is_holiday(day_dt):
            continue
        for c in CRENEAUX_JOUR:
            slot_key = f"{semaine_label}-{jour}-{c}"
            for fd in schedule_data.values():
                slot = fd['slots'].get(slot_key)
                if slot and slot[0] == groupe:
                    heures += SLOT_DURATIONS.get(c, 0)
                    break
    return heures

def get_available_salles(resolved_schedule, all_salles, semaine_label, jour, creneau):
    if not all_salles:
        return []
    slot_key = f"{semaine_label}-{jour}-{creneau}"
    occ = set()
    for fdata in resolved_schedule.values():
        slot = fdata['slots'].get(slot_key)
        if slot:
            s = slot[1]
            if "CONFLIT NON RESOLU" not in s:
                occ.add(s)
    return sorted(list(set(all_salles) - occ))

# --- SIDEBAR: Upload & processing ---
with st.sidebar:
    logo_src = get_logo_src()
    st.image(logo_src, width=200)
    st.markdown("---")
    st.markdown("### 📤 Import du Fichier (optionnel)")
    uploaded_file = st.file_uploader("Fichier Excel multi-onglets", type=['xlsx','xls'], accept_multiple_files=False, help="Sélectionnez le fichier contenant les onglets 'Planning_Mois'")
    st.text_input("Niveau (valeur export)", key="niveau_global", help="Valeur affichée dans 'Niveau' sur les exports (ex: 1ère Année)")
    st.checkbox("Activer règle 25h -> 26h (masse horaire statutaire)", value=st.session_state['force_25_to_26'], key="force_25_to_26", help="Si coché, toute masse horaire calculée à 25.0 sera remplacée par 26.0 sur les exports formateur.")
    st.markdown("---")
    st.info(f"📅 {datetime.now().strftime('%d/%m/%Y')}\n🎓 Année 2025-2026")

    if uploaded_file:
        if st.session_state['raw_data'] is None or uploaded_file != st.session_state.get('uploaded_file_ref'):
            with st.spinner("Analyse et résolution des conflits..."):
                st.session_state['raw_data'] = process_uploaded_excel(uploaded_file)
                st.session_state['uploaded_file_ref'] = uploaded_file
                if st.session_state['raw_data']:
                    st.session_state['resolved_data'], st.session_state['conflits_log'] = resolve_salle_conflits(st.session_state['raw_data'])
                    st.success(f"✅ {len(st.session_state['resolved_data'])} mois chargés et conflits traités")
                else:
                    st.session_state['resolved_data'] = None
                    st.session_state['conflits_log'] = pd.DataFrame()
                    st.error("❌ Aucune donnée valide détectée dans le fichier.")

# --- MAIN UI header ---
st.markdown(f"""
<div class="main-header">
    <div class="ofppt-title">OFPPT — Gestionnaire d'Emploi du Temps</div>
    <div class="ofppt-subtitle">Mode libre : mois et semaines saisis par l'utilisateur</div>
    <div style="font-size:0.9rem;margin-top:0.6rem;">CFP TLRA/IFMLT — Développé par ISMAILI ALAOUI Mohamed</div>
</div>
""", unsafe_allow_html=True)

resolved = st.session_state.get('resolved_data') or {}

st.markdown('<div class="section-header">⚙️ Sélection du Mois et des Semaines (Mode neutre)</div>', unsafe_allow_html=True)

col1, col2, col3 = st.columns([2,2,2])

with col1:
    imported_months = list(resolved.keys())
    imported_choice_options = ["Aucun (mode libre)"] + imported_months if imported_months else ["Aucun (mode libre)"]
    imported_month_choice = st.selectbox("Mois importé (optionnel)", imported_choice_options, index=0)
    use_imported = imported_month_choice != "Aucun (mode libre)"

with col2:
    month_label_input = st.text_input("Mois (libre) — saisissez la dénomination souhaitée", value="")
    st.caption("Ex: 'Novembre 2025' ou 'Module X - Décembre' (laisser vide si vous voulez utiliser le mois importé)")

with col3:
    st.write(" ")
    st.write(" ")
    if use_imported:
        st.info(f"Mode mixte : données importées depuis '{imported_month_choice}'. Aucun choix présélectionné.")
    else:
        st.info("Mode libre : aucun mois importé sélectionné. Saisissez un libellé de mois ci-dessus.")

# determine active month label (priority: free input > imported > empty)
if month_label_input.strip():
    active_month_label = month_label_input.strip()
elif use_imported:
    # Do NOT auto-select to comply with 'neutre' requirement: show imported choice but allow the user to keep month_label empty.
    active_month_label = imported_month_choice  # we still set active label to allow importing weeks; UI will not preselect elsewhere
else:
    active_month_label = ""

st.markdown(f"**Mois actif :** {active_month_label or '— aucun — (saisissez un mois)'}")

# initialize custom week ranges dict for active month if missing
if active_month_label and active_month_label not in st.session_state['custom_week_ranges']:
    st.session_state['custom_week_ranges'][active_month_label] = {}

# If user selected an imported month that contains week_ranges, show option to import them into the active month's custom ranges
if use_imported and active_month_label:
    imported_parsed = resolved.get(imported_month_choice)
    if imported_parsed and imported_parsed.get('week_ranges'):
        st.markdown("**Semaines détectées dans l'onglet importé :**")
        for k, v in imported_parsed['week_ranges'].items():
            st.write(f"- {k} : {v['start'].strftime('%d/%m/%Y')} → {v['end'].strftime('%d/%m/%Y')}")
        if st.button("Importer ces semaines dans le mois actif"):
            st.session_state['custom_week_ranges'][active_month_label] = {
                k: {'start': v['start'], 'end': v['end']} for k, v in imported_parsed['week_ranges'].items()
            }
            st.success(f"Semaines importées dans '{active_month_label}'.")
            st.experimental_rerun()
    else:
        st.info("Aucune plage de semaines détectée dans l'onglet importé (vous pouvez créer des semaines manuellement).")

# --- Week creation UI ---
st.markdown('<div class="section-header">🗓️ Gestion des Semaines (créez votre proposition neutre)</div>', unsafe_allow_html=True)

if not active_month_label:
    st.warning("Saisissez d'abord un libellé pour le mois (champ 'Mois (libre)') ou sélectionnez un mois importé pour activer la création de semaines.")
else:
    wk_col1, wk_col2, wk_col3, wk_col4 = st.columns([2,2,1,1])
    with wk_col1:
        new_week_label = st.text_input("Label semaine (ex: S1)", value="S1", key=f"new_week_label_{active_month_label}")
    with wk_col2:
        today = date.today()
        new_start = st.date_input("Début (lundi recommandé)", value=today, key=f"new_week_start_{active_month_label}")
        new_end = st.date_input("Fin (samedi recommandé)", value=new_start + timedelta(days=5), key=f"new_week_end_{active_month_label}")
    with wk_col3:
        if st.button("Ajouter / Mettre à jour semaine", key=f"btn_add_week_{active_month_label}"):
            label = str(new_week_label).strip() or f"S{len(st.session_state['custom_week_ranges'].get(active_month_label, {}))+1}"
            if new_end < new_start:
                st.error("La date de fin doit être postérieure ou égale à la date de début.")
            else:
                st.session_state['custom_week_ranges'].setdefault(active_month_label, {})
                st.session_state['custom_week_ranges'][active_month_label][label] = {'start': new_start, 'end': new_end}
                st.success(f"Semaine {label} enregistrée pour '{active_month_label}'.")
                st.experimental_rerun()
    with wk_col4:
        if st.button("Vider toutes les semaines du mois", key=f"btn_clear_weeks_{active_month_label}"):
            if active_month_label in st.session_state['custom_week_ranges']:
                st.session_state['custom_week_ranges'][active_month_label] = {}
                st.success(f"Toutes les semaines pour '{active_month_label}' ont été effacées.")
                st.experimental_rerun()

    # Display current weeks
    st.markdown("#### Semaines actuellement définies pour le mois actif")
    cw = st.session_state['custom_week_ranges'].get(active_month_label, {})
    if not cw:
        st.info("Aucune semaine définie. Ajoutez-en une à gauche ou importez depuis l'onglet importé.")
    else:
        rows = []
        for k, v in cw.items():
            rows.append({'Label': k, 'Début': v['start'].strftime('%d/%m/%Y'), 'Fin': v['end'].strftime('%d/%m/%Y')})
        df_cw = pd.DataFrame(rows)
        st.dataframe(df_cw, use_container_width=True)

        st.markdown("Supprimer une semaine :")
        del_label = st.selectbox("Sélectionner une semaine à supprimer", options=["--"] + sorted(list(cw.keys())), index=0, key=f"del_label_select_{active_month_label}")
        if del_label and del_label != "--":
            if st.button("Supprimer la semaine sélectionnée", key=f"btn_del_week_{active_month_label}"):
                st.session_state['custom_week_ranges'][active_month_label].pop(del_label, None)
                st.success(f"Semaine {del_label} supprimée.")
                st.experimental_rerun()

# --- Selection of active semaine (no default) ---
st.markdown('<div class="section-header">🔎 Choix de la semaine active (aucune présélection)</div>', unsafe_allow_html=True)

selected_semaine = None
active_week_ranges = st.session_state['custom_week_ranges'].get(active_month_label, {}) if active_month_label else {}
if not active_month_label:
    st.info("Aucune semaine possible tant que le mois actif n'est pas défini.")
else:
    def format_week_display_label_value(x):
        if x == "-- Aucune --":
            return x
        if not x:
            return "-- Aucune --"
        if x in active_week_ranges:
            s = active_week_ranges[x]['start']
            e = active_week_ranges[x]['end']
            return f"{x} — {s.strftime('%d/%m/%Y')} → {e.strftime('%d/%m/%Y')}"
        return x
    week_options = ["-- Aucune --"] + sorted(list(active_week_ranges.keys()), key=lambda x: int(re.sub(r'\D', '', x) or 0))
    sel = st.selectbox("Semaine (choisissez dans les semaines définies ci-dessus)", options=week_options, index=0, format_func=lambda x: format_week_display_label_value(x))
    if sel and sel != "-- Aucune --":
        selected_semaine = sel
    else:
        selected_semaine = None

# --- Inform user about imported data linkage ---
parsed = resolved.get(imported_month_choice) if use_imported else None

st.markdown('---')

if not parsed:
    st.warning("Aucune donnée importée pour le mois actif. En mode neutre, vous pouvez définir mois/semaines librement mais il n'y a pas d'emplois (importez un fichier pour afficher/exporter des plannings).")
else:
    st.success(f"Données importées pour le mois : {imported_month_choice}. Utilisez un libellé libre pour l'export si souhaité.")

# --- Tabs ---
tab1, tab2, tab3, tab4, tab5 = st.tabs(["👥 Formateurs","📚 Groupes","🚪 Salles & Conflits","📊 Salles Libres Semaine","📈 Charge par Groupe"])

with tab1:
    st.markdown('<div class="section-header">👥 Consultation / Export par Formateur</div>', unsafe_allow_html=True)
    if not parsed:
        st.info("Aucune donnée importée à afficher ici. Importez un fichier et sélectionnez le mois importé pour utiliser cette section.")
    else:
        selected_form = st.selectbox("Sélectionner un formateur", parsed['formateurs'], key="ui_form_lib")
        if selected_form and selected_semaine:
            fdata = parsed['schedule'][selected_form]
            # build table
            rows = []
            week_start = get_week_start_from_label(active_month_label or imported_month_choice, selected_semaine, active_week_ranges or parsed.get('week_ranges', {}))
            for i, jour in enumerate(JOURS):
                d = day_date(week_start, i) if week_start else None
                holiday = is_holiday(d) if d else None
                row = {'JOUR': jour}
                for c in CRENEAUX_JOUR:
                    key = f"{selected_semaine}-{jour}-{c}"
                    if holiday:
                        row[c] = ""
                    else:
                        slot = fdata['slots'].get(key, ('',''))
                        grp, salle = slot
                        row[c] = f"{grp}\n{salle}" if grp and salle else ""
                rows.append(row)
            st.dataframe(pd.DataFrame(rows), use_container_width=True)

            heures_calc = compute_hours_for_formateur(fdata, selected_semaine, active_month_label or imported_month_choice, active_week_ranges or parsed.get('week_ranges', {}))
            if st.session_state.get('force_25_to_26', True) and abs(heures_calc - 25.0) < 0.01:
                heures_display = 26.0
            else:
                heures_display = heures_calc
            st.metric("Heures (hors fériés)", f"{heures_display:.2f}h")

            st.markdown("### 📄 Export Excel")
            if st.button("📥 Générer Excel (Formateur)"):
                wb = create_excel_formateur_semaine(selected_form, fdata, selected_semaine, active_month_label or imported_month_choice, active_week_ranges or parsed.get('week_ranges', {}), niveau=st.session_state.get('niveau_global','1ère Année'), force_25_to_26=st.session_state.get('force_25_to_26', True))
                filename = sanitize_sheet_title(f"EDT_Formateur_{selected_form}_{active_month_label or imported_month_choice}", max_len=80) + ".xlsx"
                st.download_button("💾 Télécharger Excel", excel_to_bytes(wb), filename)
        else:
            st.info("Sélectionnez d'abord une semaine définie et un formateur.")

        st.markdown("---")
        if parsed and st.button("📥 Générer Pack Excel (Tous les formateurs)"):
            with st.spinner("Génération pack..."):
                wb_final = openpyxl.Workbook()
                wb_final.remove(wb_final.active)
                used_names = set()
                for form in parsed['formateurs']:
                    wb_temp = create_excel_formateur_semaine(form, parsed['schedule'][form], selected_semaine or (parsed['semaines'][0] if parsed['semaines'] else None), active_month_label or imported_month_choice, active_week_ranges or parsed.get('week_ranges', {}), niveau=st.session_state.get('niveau_global','1ère Année'), force_25_to_26=st.session_state.get('force_25_to_26', True))
                    ws_temp = wb_temp.active
                    sheet_base = sanitize_sheet_title(f"{form[:25]}_{active_month_label or imported_month_choice}", max_len=31)
                    sheet_name = sheet_base
                    i = 1
                    while sheet_name in used_names:
                        suffix = f"_{i}"
                        sheet_name = sanitize_sheet_title(sheet_base[:31-len(suffix)] + suffix)
                        i += 1
                    used_names.add(sheet_name)
                    ws_new = wb_final.create_sheet(title=sheet_name)
                    copy_sheet(ws_temp, ws_new)
                filename = sanitize_sheet_title(f"Pack_Formateurs_{active_month_label or imported_month_choice}", max_len=80) + ".xlsx"
                st.download_button("💾 Télécharger Pack Excel (Formateurs)", excel_to_bytes(wb_final), filename)

with tab2:
    st.markdown('<div class="section-header">📚 Consultation / Export par Groupe</div>', unsafe_allow_html=True)
    if not parsed:
        st.info("Aucune donnée importée à afficher ici.")
    else:
        selected_grp = st.selectbox("Sélectionner un groupe", parsed['groupes'], key="ui_grp_lib")
        if selected_grp and selected_semaine:
            # build groupe table
            rows = []
            week_start = get_week_start_from_label(active_month_label or imported_month_choice, selected_semaine, active_week_ranges or parsed.get('week_ranges', {}))
            for i, jour in enumerate(JOURS):
                d = day_date(week_start, i) if week_start else None
                holiday = is_holiday(d) if d else None
                row = {'JOUR': jour}
                for c in CRENEAUX_JOUR:
                    key = f"{selected_semaine}-{jour}-{c}"
                    info = ""
                    for f, fd in parsed['schedule'].items():
                        s = fd['slots'].get(key)
                        if s and s[0] == selected_grp:
                            info = f"{f}\n{s[1].replace(' (CONFLIT NON RESOLU)',' (Conflit)')}"
                            break
                    row[c] = info
                rows.append(row)
            st.dataframe(pd.DataFrame(rows), use_container_width=True)
            heures_g = compute_hours_for_groupe(parsed['schedule'], selected_grp, selected_semaine, active_month_label or imported_month_choice, active_week_ranges or parsed.get('week_ranges', {}))
            st.metric("Heures (hors fériés)", f"{heures_g:.2f}h")

            if st.button("📥 Générer Excel (Groupe)"):
                wb = create_excel_groupe_semaine(selected_grp, parsed['schedule'], selected_semaine, active_month_label or imported_month_choice, active_week_ranges or parsed.get('week_ranges', {}), niveau=st.session_state.get('niveau_global','1ère Année'))
                filename = sanitize_sheet_title(f"EDT_Groupe_{selected_grp}_{active_month_label or imported_month_choice}", max_len=80) + ".xlsx"
                st.download_button("💾 Télécharger Excel", excel_to_bytes(wb), filename)
        else:
            st.info("Sélectionnez d'abord une semaine définie et un groupe.")

        st.markdown("---")
        if parsed and st.button("📥 Générer Pack Excel (Tous les groupes)"):
            with st.spinner("Génération pack..."):
                wb_final = openpyxl.Workbook()
                wb_final.remove(wb_final.active)
                used_names = set()
                for groupe in parsed['groupes']:
                    wb_temp = create_excel_groupe_semaine(groupe, parsed['schedule'], selected_semaine or (parsed['semaines'][0] if parsed['semaines'] else None), active_month_label or imported_month_choice, active_week_ranges or parsed.get('week_ranges', {}), niveau=st.session_state.get('niveau_global','1ère Année'))
                    ws_temp = wb_temp.active
                    sheet_base = sanitize_sheet_title(f"{groupe[:25]}_{active_month_label or imported_month_choice}", max_len=31)
                    sheet_name = sheet_base
                    i = 1
                    while sheet_name in used_names:
                        suffix = f"_{i}"
                        sheet_name = sanitize_sheet_title(sheet_base[:31-len(suffix)] + suffix)
                        i += 1
                    used_names.add(sheet_name)
                    ws_new = wb_final.create_sheet(title=sheet_name)
                    copy_sheet(ws_temp, ws_new)
                filename = sanitize_sheet_title(f"Pack_Groupes_{active_month_label or imported_month_choice}", max_len=80) + ".xlsx"
                st.download_button("💾 Télécharger Pack Excel (Groupes)", excel_to_bytes(wb_final), filename)

with tab3:
    st.markdown('<div class="section-header">🚪 Salles & Conflits</div>', unsafe_allow_html=True)
    colj, colc, cold = st.columns(3)
    with colj:
        sel_jour = st.selectbox("Jour", JOURS, key="salle_jour")
    with colc:
        sel_cr = st.selectbox("Créneau", CRENEAUX_JOUR, key="salle_cr")
    if not parsed:
        st.info("Aucune donnée importée — pas de conflits détectables en mode neutre.")
    else:
        salles_libres = get_available_salles(parsed['schedule'], parsed['salles'], selected_semaine or (parsed['semaines'][0] if parsed['semaines'] else ''), sel_jour, sel_cr) if sel_jour and sel_cr else []
        st.metric("Salles disponibles", len(salles_libres))
        if salles_libres:
            st.write(", ".join(salles_libres))
        else:
            st.write("Aucune salle disponible")

        st.markdown("---")
        conflits = st.session_state['conflits_log']
        if conflits.empty:
            st.info("Aucun conflit détecté.")
        else:
            cs = conflits
            if use_imported and imported_month_choice:
                cs = conflits[(conflits['Mois']==imported_month_choice) & (conflits['Semaine']==selected_semaine)] if selected_semaine else conflits[conflits['Mois']==imported_month_choice]
            st.dataframe(cs, use_container_width=True)
            if not cs.empty:
                b = BytesIO()
                cs.to_excel(b, index=False, sheet_name='Conflits')
                b.seek(0)
                st.download_button("📥 Télécharger Conflits", b.getvalue(), f"Conflits_{imported_month_choice or 'ALL'}_{selected_semaine or 'ALL'}.xlsx")

with tab4:
    st.markdown('<div class="section-header">📊 Synthèse Salles Libres</div>', unsafe_allow_html=True)
    if not parsed:
        st.info("Aucune donnée importée : synthèse non disponible.")
    else:
        if not selected_semaine:
            st.info("Sélectionnez une semaine définie pour obtenir la synthèse des salles libres.")
        else:
            synth = []
            week_start = get_week_start_from_label(active_month_label or imported_month_choice, selected_semaine, active_week_ranges or parsed.get('week_ranges', {}))
            for jour in JOURS:
                for c in CRENEAUX_JOUR:
                    key = f"{selected_semaine}-{jour}-{c}"
                    d = day_date(week_start, JOURS.index(jour))
                    holiday = True if (d and is_holiday(d)) else False
                    occ = set()
                    if not holiday:
                        for f, fd in parsed['schedule'].items():
                            s = fd['slots'].get(key)
                            if s and s[0]:
                                occ.add(s[1].replace(' (CONFLIT NON RESOLU)','').replace(' (Conflit)',''))
                    libres = sorted(list(set(parsed['salles']) - occ))
                    synth.append({'Jour': jour, 'Créneau': c, 'Horaire': HORAIRES[c], 'Nb Salles Libres': len(libres), 'Salles Disponibles': ', '.join(libres) if libres else 'Aucune'})
            st.dataframe(pd.DataFrame(synth), use_container_width=True)

with tab5:
    st.markdown('<div class="section-header">📈 Charge par Groupe</div>', unsafe_allow_html=True)
    if not parsed:
        st.info("Aucune donnée importée : analyse de charge non disponible.")
    else:
        if not selected_semaine:
            st.info("Sélectionnez une semaine définie pour effectuer l'analyse de charge.")
        else:
            charge_groupes = []
            for groupe in parsed['groupes']:
                heures_total = 0
                nb_creneaux = 0
                for jour in JOURS:
                    for creneau in CRENEAUX_JOUR:
                        slot_key = f"{selected_semaine}-{jour}-{creneau}"
                        for formateur, f_data in parsed['schedule'].items():
                            slot_data = f_data['slots'].get(slot_key)
                            if slot_data and slot_data[0] == groupe:
                                heures_total += SLOT_DURATIONS[creneau]
                                nb_creneaux += 1
                                break
                charge_groupes.append({'Groupe': groupe, 'Heures de Formation': heures_total, 'Nombre de Créneaux': nb_creneaux})

            if not charge_groupes:
                st.info("Aucune donnée de charge disponible pour la semaine sélectionnée.")
            else:
                df_charge = pd.DataFrame(charge_groupes).sort_values('Heures de Formation', ascending=False)
                moyenne_heures = df_charge['Heures de Formation'].mean()
                col_met1, col_met2, col_met3, col_met4 = st.columns(4)
                with col_met1:
                    st.metric("Groupes Total", len(df_charge))
                with col_met2:
                    st.metric("Charge Moyenne", f"{moyenne_heures:.1f}h")
                with col_met3:
                    st.metric("Charge Minimale", f"{df_charge['Heures de Formation'].min():.1f}h")
                with col_met4:
                    st.metric("Charge Maximale", f"{df_charge['Heures de Formation'].max():.1f}h")

                import plotly.graph_objects as go
                seuil_bas = moyenne_heures * 0.85
                seuil_haut = moyenne_heures * 1.15
                colors = []
                for heures in df_charge['Heures de Formation']:
                    if heures > seuil_haut:
                        colors.append('#d32f2f')
                    elif heures >= seuil_bas and heures <= seuil_haut:
                        colors.append('#fbc02d')
                    else:
                        colors.append('#388e3c')
                fig = go.Figure(data=[go.Bar(x=df_charge['Groupe'], y=df_charge['Heures de Formation'], text=df_charge['Heures de Formation'].apply(lambda x: f"{x:.1f}h"), textposition='outside', marker=dict(color=colors, line=dict(color='#1e5631', width=1.5)), hovertemplate='<b>%{x}</b><br>Heures: %{y:.1f}h<br><extra></extra>')])
                fig.add_hline(y=moyenne_heures, line_dash="dash", line_color="#1e5631", annotation_text=f"Moyenne: {moyenne_heures:.1f}h", annotation_position="right")
                fig.add_hline(y=seuil_haut, line_dash="dot", line_color="#d32f2f", opacity=0.5)
                fig.add_hline(y=seuil_bas, line_dash="dot", line_color="#388e3c", opacity=0.5)
                fig.update_layout(title={'text': f'Charge Horaire par Groupe - {active_month_label or imported_month_choice} {selected_semaine}', 'x': 0.5}, xaxis_title='Groupes', yaxis_title='Heures de Formation', plot_bgcolor='white', paper_bgcolor='#f8faf9', height=500, showlegend=False, xaxis=dict(tickangle=-45, gridcolor='lightgray'), yaxis=dict(gridcolor='lightgray'))
                st.plotly_chart(fig, use_container_width=True)

                df_charge['Catégorie'] = df_charge['Heures de Formation'].apply(lambda x: "🔴 Trop Chargé" if x > seuil_haut else ("🟡 Chargé" if x >= seuil_bas else "🟢 Normal"))
                df_charge['Écart/Moyenne'] = (df_charge['Heures de Formation'] - moyenne_heures).apply(lambda x: f"{x:+.1f}h")
                st.dataframe(df_charge, use_container_width=True)

                st.markdown("---")
                if st.button("📥 Exporter l'Analyse de Charge (Excel)"):
                    wb_charge = openpyxl.Workbook()
                    ws = wb_charge.active
                    ws.title = sanitize_sheet_title("Charge_Groupes")
                    ws.sheet_view.showGridLines = False
                    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    header_font = Font(bold=True, size=11, color="FFFFFF")
                    title_font = Font(bold=True, size=14, color="1e5631")
                    header_fill = PatternFill(start_color="2d8659", end_color="2d8659", fill_type="solid")
                    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws['A1'] = f'ANALYSE DE CHARGE PAR GROUPE - {active_month_label or imported_month_choice} {selected_semaine}'
                    ws.merge_cells('A1:E1'); ws['A1'].font = title_font; ws['A1'].alignment = center_align; ws.row_dimensions[1].height = 25
                    ws['A2'] = f'Moyenne: {moyenne_heures:.1f}h | Seuils: Normal < {seuil_bas:.1f}h | Chargé: {seuil_bas:.1f}h-{seuil_haut:.1f}h | Trop Chargé > {seuil_haut:.1f}h'
                    ws.merge_cells('A2:E2'); ws['A2'].alignment = center_align; ws.row_dimensions[2].height = 20
                    ws['A4'] = 'Groupe'; ws['B4'] = 'Heures de Formation'; ws['C4'] = 'Nombre de Créneaux'; ws['D4'] = 'Niveau de Charge'; ws['E4'] = 'Écart/Moyenne'
                    for col in ['A','B','C','D','E']:
                        ws[f'{col}4'].font = header_font; ws[f'{col}4'].fill = header_fill; ws[f'{col}4'].border = border_thin; ws[f'{col}4'].alignment = center_align; ws.column_dimensions[col].width = 25
                    row = 5
                    for _, data_row in df_charge.iterrows():
                        ws[f'A{row}'] = data_row['Groupe']; ws[f'B{row}'] = data_row['Heures de Formation']; ws[f'C{row}'] = data_row['Nombre de Créneaux']; ws[f'D{row}'] = data_row['Catégorie']; ws[f'E{row}'] = data_row['Écart/Moyenne']
                        for col in ['A','B','C','D','E']:
                            ws[f'{col}{row}'].border = border_thin; ws[f'{col}{row}'].alignment = center_align
                        row += 1
                    excel_bytes = excel_to_bytes(wb_charge)
                    st.download_button("💾 Télécharger l'Analyse", excel_bytes, f"Charge_Groupes_{active_month_label or imported_month_choice}_{selected_semaine}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

st.markdown("---")
st.markdown("<div style='text-align:center;color:#666;padding:1rem;'>Mode libre — mois et semaines neutres. Importez un fichier pour lier des données réelles.</div>", unsafe_allow_html=True)
